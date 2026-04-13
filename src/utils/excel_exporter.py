# -*- coding: utf-8 -*-
"""Excel导出工具"""
import os
from datetime import datetime
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

class ExcelExporter:
    """Excel导出器"""
    
    @staticmethod
    def export(data, filepath, headers=None):
        """
        导出数据到Excel
        data: 二维列表 [[row1], [row2], ...]
        headers: 表头列表
        """
        if not HAS_OPENPYXL:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DICOM文件列表"
        
        # 写入表头
        if headers:
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        # 保存
        wb.save(filepath)
        return True
